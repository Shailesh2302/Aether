from typing import List
from openpyxl import load_workbook
from app.core.logger import app_logger


class ExcelLoader:
    async def load_file(self, file_path: str) -> List[str]:
        app_logger.info(f"Loading Excel: {file_path}")
        wb = load_workbook(file_path, read_only=True, data_only=True)
        sheets = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows = []
            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join(str(c) if c is not None else "" for c in row)
                if row_text.strip():
                    rows.append(row_text)
            if rows:
                sheets.append(f"Sheet: {sheet_name}\n" + "\n".join(rows))
        wb.close()
        return sheets if sheets else [""]
